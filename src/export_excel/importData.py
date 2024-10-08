# export.py
import sys
sys.path.append("") 

import openpyxl
import os
import shutil
from datetime import datetime

from src.export_excel.config import (input_1_excel, input_2_excel, output_2_excel,
                                    input_3_exel, FORMAT_NUMBER)

def clear_sheet(src_file: str, dest_file: str, sheet_name: str):
    """
    Replaces a sheet in the destination Excel file with a sheet from the source Excel file. 
    Clear the current sheet by copying the sheet from template
    
    :param src_file: Path to the source Excel file.
    :param dest_file: Path to the destination Excel file.
    :param sheet_name: The name of the sheet to copy from the source file and replace in the destination file.
    """
    try:
        # Load the source workbook and get the sheet to copy
        src_wb = openpyxl.load_workbook(src_file)
        if sheet_name not in src_wb.sheetnames:
            return
        src_sheet = src_wb[sheet_name]

        # Load the destination workbook
        dest_wb = openpyxl.load_workbook(dest_file)

        # If the sheet already exists in the destination, remove it
        if sheet_name in dest_wb.sheetnames:
            dest_wb.remove(dest_wb[sheet_name])

        # Create a new sheet in the destination workbook and copy the contents
        dest_sheet = dest_wb.create_sheet(title=sheet_name)

        # Copy the values from the source sheet to the destination sheet
        for row in src_sheet.iter_rows(values_only=True):
            dest_sheet.append(row)

        # Save the destination workbook with the updated sheet
        dest_wb.save(dest_file)

    except Exception as e:
        print(f"Error occurred: {e}")

def choose_sheet_to_write(excel_path: str, invoice_1: dict):
    """
    Choose sheet name 'Vorgang' in VorlageSpesenabrechnung.xlsx to write data to.
    Returns a tuple (sheet_name, is_modified) where is_modified is True if modifying an existing sheet,
    and False if writing data to a new (empty) sheet.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_path)
    
    project_number = invoice_1.get('project_number', '').strip().lower()
    
    # Check existing sheets to see if we are modifying a sheet
    for i in range(1, 6):
        sheet_name = f'Vorgang {i}'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            cell_value = sheet['B4'].value
            
            # Check if project number matches
            if cell_value and str(cell_value).lower() == project_number:
                # Return sheet name and flag as True (sheet will be modified)
                return sheet_name, True
    
    # If no match is found, find the first empty sheet
    for i in range(1, 6):
        sheet_name = f'Vorgang {i}'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            cell_value = sheet['B4'].value
            
            # Check if the sheet is empty (B4 is None or empty)
            if cell_value is None or cell_value == '':
                # Return sheet name and flag as False (writing to a new sheet)
                return sheet_name, False
    
    # If all sheets are full, return None for the sheet name and False for the flag
    return None, False


def prepare_excel_files(employee_expense_report_path: str, output_path: str = "output"):
    """
    Prepare Excel files in the specified output directory.
    
    - Ensure the output directory exists.
    - Copy the input_2_excel file to output_2_excel if not already present.
    - If employee_expense_report_path does not exist, copy input_1_excel to output_path and rename it.
    """

    # Ensure output_path exists; if not, create it
    if not os.path.exists(output_path):
        os.makedirs(output_path)
        print(f"Created output directory: {output_path}")
    
    # Check if output_2_excel exists, if not, copy from input_2_excel
    if not os.path.exists(output_2_excel):
        try:
            shutil.copy(input_2_excel, output_2_excel)
            print(f"Copied: {input_2_excel} to {output_2_excel}")
        except Exception as e:
            print(f"Error when copying {input_2_excel} to {output_2_excel}: {e}")
    
    # Check if employee_expense_report_path exists, if not, copy input_1_excel to output_path and rename it
    if not os.path.exists(employee_expense_report_path):
        try:
            # Define the destination path for the employee expense report
            dest_employee_path = os.path.join(output_path, os.path.basename(employee_expense_report_path))
            
            # Copy the input_1_excel to the destination and rename it
            shutil.copy(input_1_excel, dest_employee_path)
            print(f"Copied and renamed: {input_1_excel} to {dest_employee_path}")
        except Exception as e:
            print(f"Error when copying {input_1_excel} to {employee_expense_report_path}: {e}")
    else:
        print(f"{employee_expense_report_path} already exists, no need to copy.")


def update_excel(json_data:dict, excel_des_path:str="output/VorlageSpesenabrechnung.xlsx", 
                 sheet_name:str="", cell:str="", value:str=""):
    '''
    Input:
        json_index: index of json - define in config.py - to set path
        excel_index: index of excel - define in config.py - to set path
        sheet_name: sheet of excel file - write data to this sheet
        cell: cell want  to write data
        value: data need to be written to cell
    '''
        
    # Access the value from the JSON
    if value == 'lines':
        handle_lines_value(json_data, sheet_name, excel_des_path)
        return
    elif value == 'cost_hour':
        handle_number_hour_value(excel_des_path, sheet_name)
        return
    elif value == 'service':
        handle_add_service(json_data, excel_des_path, sheet_name)
        return
    elif value in json_data:
        value_to_write = json_data[value]
        # Write the existing Excel workbook
        write_data(excel_des_path, sheet_name, cell, value_to_write)
    else:
        print(f"Key '{value}' not found in the JSON data.")
        return

def write_data(excel_des_path, sheet_name, cell, value_to_write, is_money = False):
    '''
    Input:
        excel_index: index of excel - define in config.py - to set path
        sheet_name: sheet of excel file - write data to this sheet
        cell: cell want  to write data
        value_to_write: data need to be written to cell
        is_money: to convert format money
    '''
    # Load the existing Excel workbook
    excel_des = excel_des_path
    
    wb = openpyxl.load_workbook(excel_des)

    # Access the specified sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"Sheet '{sheet_name}' not found in the Excel file.")
        return
    
    # Write the value to the specified cell
    ws[cell].value = value_to_write
    if is_money == True:
        ws[cell].number_format = FORMAT_NUMBER

    # Save the updated Excel workbook
    wb.save(excel_des)
    wb.close()

def conver_number_hour(time):
    '''
    Hanlde time
    '''
    # Split the time into hours, minutes, and seconds
    hours, minutes, seconds = map(int, time.split(':'))

    # Convert the time into a float (hours + minutes/60)
    return float(hours + minutes / 60 + seconds / 3600)


def handle_lines_value(json_data, sheet_name, excel_des_path):
    '''
    json_data: file json 2
    sheet_name: sheet want to write data
    excel_index: index excel - config at config.py
    '''
    number = get_last_row(excel_des_path, sheet_name, 'A', 11)
    for line in json_data['lines']:
        cell_des = f'A{number}'
        value_to_write = line['date']
        date_obj = datetime.strptime(value_to_write, '%Y-%m-%d')
        formatted_date = date_obj.strftime('%d.%m.%Y')
        write_data(excel_des_path, sheet_name, cell_des, formatted_date)

        cell_des = f'B{number}'
        value_to_write = line['start_time']
        write_data(excel_des_path, sheet_name, cell_des, conver_number_hour(value_to_write))

        cell_des = f'C{number}'
        value_to_write = line['end_time']
        write_data(excel_des_path, sheet_name, cell_des, conver_number_hour(value_to_write))

        number += 1

def handle_number_hour_value(excel_des_path, sheet_name):
    '''
    sheet_name: sheet want to write data
    excel_des_path: path to the destination Excel file
    cell: cell to write data (not used in this function)
    '''
    # Load the source workbook and sheet
    source_workbook = openpyxl.load_workbook(input_3_exel)
    source_sheet = source_workbook['Table 1']
    
    # Load the target workbook and sheet
    target_workbook = openpyxl.load_workbook(excel_des_path)
    target_sheet = target_workbook[sheet_name]

    # Find land and city
    land = target_sheet['B8'].value
    city = target_sheet['B6'].value

    values_to_write = None

    for row in range(1, source_sheet.max_row + 1):
        current_land = source_sheet[f'A{row}'].value
        
        if current_land == land:
            # If land matches and has values, use these values
            if source_sheet[f'B{row}'].value:
                values_to_write = [source_sheet[f'B{row}'].value, 
                                   source_sheet[f'C{row}'].value, 
                                   source_sheet[f'D{row}'].value]
                break
            # If land matches but has no values, look for city or "im Übrigen"
            else:
                for sub_row in range(row + 1, source_sheet.max_row + 1):
                    sub_entry = source_sheet[f'A{sub_row}'].value
                    if sub_entry and sub_entry.strip().startswith('–'):
                        if city in sub_entry:
                            values_to_write = [source_sheet[f'B{sub_row}'].value, 
                                               source_sheet[f'C{sub_row}'].value, 
                                               source_sheet[f'D{sub_row}'].value]
                            break
                        elif "im Übrigen" in sub_entry:
                            values_to_write = [source_sheet[f'B{sub_row}'].value, 
                                               source_sheet[f'C{sub_row}'].value, 
                                               source_sheet[f'D{sub_row}'].value]
                    elif sub_entry and not sub_entry.strip().startswith('–'):
                        break  # We've reached the next main entry
                break  # Exit the outer loop as we've processed this land

    source_workbook.close()
    target_workbook.close()

    if values_to_write:
        write_data(excel_des_path, sheet_name, 'F2', values_to_write[0], True)
        write_data(excel_des_path, sheet_name, 'G2', values_to_write[1], True)
        write_data(excel_des_path, sheet_name, 'H2', values_to_write[2], True)
    else:
        print(f"No matching data found for land: {land}, city: {city}")


def handle_add_service(json_data, excel_des_path, sheet_name):
    date = json_data['sign_date']
    date_obj = datetime.strptime(date, '%Y-%m-%d')
    date = date_obj.strftime('%d.%m.%Y')

    line = get_line_from_date(excel_des_path, sheet_name, date)

    # Handle with hotel
    for fixed_line in json_data['fixed_lines']:
        if fixed_line['title'] == 'Hotel':
            if fixed_line['payment_method'] == 'self paid':
                value_to_write = fixed_line['amount']
                cell = f'E{line}'
                write_data(excel_des_path, sheet_name, cell, value_to_write, True)
            elif fixed_line['with_breakfast'] == True:
                value_to_write = get_breakfas_value(excel_des_path, sheet_name)
                cell = f'F{line}'
                write_data(excel_des_path, sheet_name, cell, value_to_write, True)
            
            fix_date_with_hotel(excel_des_path, sheet_name, line)
            break

    # Handle with lines
    value = 0
    title = ''
    last_line = get_last_row(excel_des_path, sheet_name, 'H', 11)

    for row in json_data['lines']:
        value = row['amount']
        title = row['title'] 

        cell = f'H{last_line}'
        write_data(excel_des_path, sheet_name, cell, value, True)
        cell = f'I{last_line}'
        write_data(excel_des_path, sheet_name, cell, title)

        last_line = last_line + 1


def get_line_from_date(excel_des_path, sheet_name, date):
    # Load the existing Excel workbook
    excel_des = excel_des_path
    wb = openpyxl.load_workbook(excel_des)
    # Access the specified sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"Sheet '{sheet_name}' not found in the Excel file.")
        return
        
    start_row = 11
    for row in range(start_row, ws.max_row + 1):
        cell = ws['A' + str(row)]
        if cell.value == date:
            return row
        
    return None

def get_breakfas_value(excel_des_path, sheet_name):
    excel_des = excel_des_path

    wb = openpyxl.load_workbook(excel_des, data_only=True)
    # Access the specified sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"Sheet '{sheet_name}' not found in the Excel file.")
        return 0

    money = ws['G2'].value * 0.2
    wb.close()
    return money

def get_last_row(excel_des_path, sheet_name, column, start_row):
    '''
    sheet_name: sheet want to write data
    excel_des_path: index excel - config at config.py
    column: column to get last row
    start_row:  row start

    Output: las row without data
    '''
    # Load the existing Excel workbook
    excel_des = excel_des_path
    wb = openpyxl.load_workbook(excel_des)
        # Access the specified sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"Sheet '{sheet_name}' not found in the Excel file.")
        return
    
    for row in range(start_row, ws.max_row + 1):
        cell = ws[column + str(row)]
        if cell.value == None:
            return row

def fix_date_with_hotel(excel_des_path, sheet_name, line):
    '''
    Fixed date if hotel enable
    '''
    last_line = get_last_row(excel_des_path, sheet_name, 'A', line)

    value_to_write = 24
    cell = f'C{line}'
    write_data(excel_des_path, sheet_name, cell, value_to_write)

    for row in range(line + 1, last_line - 1):
        value_to_write = 0
        cell = f'B{row}'
        write_data(excel_des_path, sheet_name, cell, value_to_write)

        value_to_write = 24
        cell = f'C{row}'
        write_data(excel_des_path, sheet_name, cell, value_to_write)

    value_to_write = 0
    cell = f'B{row + 1}'
    write_data(excel_des_path, sheet_name, cell, value_to_write)