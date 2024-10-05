import sys
sys.path.append("")

import os
import formulas
import shutil

import openpyxl
from openpyxl import load_workbook

from src.export_excel.config import output_1_excel, output_2_excel

def copy_data_individual_to_group(src_file_path:str = "output/StDi_07_24.xlsx", 
                                  src_sheet:str = "Auszahlung", 
                                  des_file_path:str = "output/1.4437_10578_A3DS GmbH_04_2024 .xlsm", 
                                  des_sheet:str = "April 24_0"):
    """Copy data in 3 column D5, E5, G5 from each employee excel to the big excel of all employess to sum"""
    

    # Load the source workbook and sheet
    source_workbook = openpyxl.load_workbook(src_file_path, data_only=True)
    source_sheet = source_workbook[src_sheet]
    
    # Load the target workbook and sheet
    target_workbook = openpyxl.load_workbook(des_file_path, keep_vba=True)
    target_sheet = target_workbook[des_sheet]

    # Get name from src file
    full_name   = source_sheet['B1'].value
    name_parts = full_name.split()
    nachname    = name_parts[0]
    vorname     = " ".join(name_parts[1:])

    # Iterate through the rows in the source file (e.g., name is in column A)
    for target_row in range(11, target_sheet.max_row + 1):
        target_nachname = target_sheet[f'E{target_row}'].value  # Name from source file (column A)
        target_vorname  = target_sheet[f'J{target_row}'].value  # Name from source file (column A)

        if target_nachname == nachname and target_vorname == vorname:
            target_sheet[f'T{target_row}'].value = get_formula_result(src_file_path, src_sheet, 'D5')
            target_sheet[f'V{target_row}'].value = get_formula_result(src_file_path, src_sheet, 'E5')
            target_sheet[f'U{target_row}'].value = get_formula_result(src_file_path, src_sheet, 'G5')
            print(f"Updated infomation of '{full_name}'")
            break 

    target_workbook.save(des_file_path)


def get_formula_result(file_path, sheet_name, cell_address):
    # Create a temporary directory
    """Read value of a formulas cell without using xlwings package"""
    temp_dir = 'temp'
    os.makedirs(temp_dir, exist_ok=True)

    # Get the base filename and create the new uppercase filename
    base_filename = os.path.basename(file_path)
    new_filename = base_filename.upper()
    temp_file_path = os.path.join(temp_dir, new_filename)

    try:
        # Use formulas to calculate and save the new version of the Excel file
        xl_model = formulas.ExcelModel().loads(file_path).finish()
        xl_model.calculate()
        xl_model.write(dirpath=temp_dir)

        # Load the workbook using openpyxl
        workbook = load_workbook(filename=temp_file_path, data_only=True)
        
        # Get the specified sheet
        sheet = workbook[sheet_name.upper()]
        
        # Get the value from the specified cell
        result = sheet[cell_address].value
        
        # Close the workbook
        workbook.close()
        
        return result

    finally:
        # Clean up: remove the temporary directory and its contents
        shutil.rmtree(temp_dir, ignore_errors=True)
        # pass


if __name__ == "__main__":
    from config import output_1_excel, output_2_excel, MAIN_SHEET

    print(get_formula_result(file_path = "output/StDi_07_24.xlsx", sheet_name=MAIN_SHEET, cell_address='D5'))
    