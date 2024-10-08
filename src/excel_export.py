import sys
sys.path.append("")

import os
import openpyxl
from typing import List, Tuple, Dict
from src.Utils.utils import read_config, find_best_match_fuzzy


class EmployeeNameRetriever:
    def __init__(self, config_path:str = None, config:dict = None):
        # Load configuration from the YAML file
        if (config_path is None and config is None) or (config_path and config):
            raise ValueError("Either 'config_path' or 'config' must be provided, but not both or neither.")

        if config_path:
            self.config_path = config_path
            self.config = read_config(path=self.config_path)['excel']
        elif config:
            self.config = config['excel']['employee_name']

        # Load the Excel file and sheet
        self.workbook = openpyxl.load_workbook(self.config['excel_file_path'], keep_vba=True)
        self.sheet = self.workbook[self.config['sheet_name']]
        self.nachname_position = None
        self.vorname_position = None
        self.canonical_names = self._preprocess_names()

    def _find_nachname_vorname_positions(self):
        """Find the row and column positions of 'nachname' and 'vorname'."""
        nachname_key = self.config['nachname_key'].lower()
        vorname_key = self.config['vorname_key'].lower()
        
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, str):
                        if cell.value.lower() == nachname_key:
                            self.nachname_position = (cell.row, cell.column)
                        elif cell.value.lower() == vorname_key:
                            self.vorname_position = (cell.row, cell.column)
                    
                if self.nachname_position and self.vorname_position:
                    break
            if self.nachname_position and self.vorname_position:
                break

        # Check if the positions are valid
        if not (self.nachname_position and self.vorname_position):
            raise ValueError(f"Either '{nachname_key}' or '{vorname_key}' not found in the sheet.")
        if self.nachname_position[1] >= self.vorname_position[1]:
            raise ValueError(f"'{nachname_key}' must be in a column less than '{vorname_key}'.")

    def get_user_names(self):
        """Returns a list of tuples (nachname, vorname) starting from the row after the header."""
        self._find_nachname_vorname_positions()  # Ensure the positions are set
        
        row_num = self.nachname_position[0] + 1
        values_nachname_and_vorname = []

        while True:
            # Get values from the nachname and vorname columns
            nachname_value = self.sheet.cell(row=row_num, column=self.nachname_position[1]).value
            vorname_value = self.sheet.cell(row=row_num, column=self.vorname_position[1]).value
            
            # Break the loop if both are None (assumes end of data)
            if nachname_value is None and vorname_value is None:
                break

            # Append the values (even if one is None)
            values_nachname_and_vorname.append((nachname_value, vorname_value))
            
            # Move to the next row
            row_num += 1

        return values_nachname_and_vorname
    
    def get_user_email(self, nachname: str, vorname: str) -> str:
        """
        Generate the user email based on the first two characters of the last and first names.
        :param nachname: Last name of the user.
        :param vorname: First name of the user.
        :return: Email address as a string.
        """
        # Get the first two characters from each name, ensuring they are not None
        nachname_part = nachname[:2].lower() if nachname else ''
        vorname_part = vorname[:2].lower() if vorname else ''
        return f"{nachname_part}{vorname_part}@gmail.com"

    def get_user_info(self):
        """
        Returns a list of dictionaries, each containing user information:
        {'name': (nachname, vorname), 'email': email}
        """
        # Get the list of user names
        user_names = self.get_user_names()
        user_info_list = []

        # Generate user info with names and emails
        for nachname, vorname in user_names:
            email = self.get_user_email(nachname, vorname)
            user_info_list.append({
                'name': (nachname, vorname),
                'email': email
            })

        return user_info_list
    
    def get_sheet_names(self):
        return self.workbook.sheetnames

    def _preprocess_names(self) -> List[Tuple[str, int]]:
        """
        Preprocess the names by creating both first-last and last-first formats.
        :return: List of unique preprocessed names as strings with their original indices.
        """
        corpus = []
        names = self.get_user_names()
        for idx, (last_name, first_name) in enumerate(names):
            last_first = f"{last_name.lower()} {first_name.lower()}"
            first_last = f"{first_name.lower()} {last_name.lower()}"
            corpus.append((last_first, idx))
            corpus.append((first_last, idx))
        return corpus

    def find_best_matching_name(self, name: str, name_thresh=None) -> Tuple[Tuple[str, str], float]:
        """
        Find the closest match to the given name using fuzzy matching.
        :param name: Input name string (possibly incorrect)
        :return: Tuple containing the best matching name (nachname, vorname) and the similarity score
        """
        name_list = [name for name, _ in self.canonical_names]
        best_idx, best_match, best_score = find_best_match_fuzzy(string_list=name_list, 
                                                           text=name)
        
        # Find the index of the best match in canonical_names
        best_idx = next(idx for idx, (n, _) in enumerate(self.canonical_names) if n == best_match)
        
        # Get the original index of the name in the user_names list
        original_idx = self.canonical_names[best_idx][1]
        
        # Get the original name tuple
        if not name_thresh:
            name_thresh = self.config['name_thresh']
        
        if best_score >= name_thresh:
            return self.get_user_names()[original_idx]
        else:
            return ""  # Normalize score to be between 0 and 1

 

def get_full_name(name_tuple):
    """
    Convert a tuple into a full name string.
    :param name_tuple: Tuple containing parts of a name (e.g., last_name, first_name, etc.)
    :return: Full name as a single string with space-separated values.
    """
    return ' '.join(f"{part}" for part in name_tuple)


if __name__ == "__main__":
    # Example usage:
    config_path = 'config/config.yaml'
    config = read_config(config_path)
    processor = EmployeeNameRetriever(config=config)
    names = processor.get_user_names()

    print(names)
    user_info = processor.get_user_info()
    for info in user_info:
        print(info)
        
    print(processor.get_sheet_names())

    # OCR output (potentially with errors and mixed order)
    ocr_output = ["Téuuley Divl", "Tuuulev Dirk", "Tümmeler Dirk", "Dirk Tuuulev", "Divl Téuuley"]

    for ocr_name in ocr_output:
        print(processor.find_best_matching_name(name = ocr_name, ))


